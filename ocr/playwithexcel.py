from openpyxl import Workbook, load_workbook

# book = Workbook()
book = load_workbook(r"C:\Users\KIRUBA\Documents\crewai_new_workflow\info.xlsx")
sheet = book.active


values = []
for i in sheet.iter_rows():
    for j in i:
        if j.value == '23': 
            values.append((j.row, j.column))
    
print(values)